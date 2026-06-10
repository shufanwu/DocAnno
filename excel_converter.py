from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import json


class ExcelConverter:
    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def save_excel(self, table_data, output_path):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        
        data = table_data.get('data', [])
        merge_cells = table_data.get('mergeCells', [])
        
        for row_idx, row in enumerate(data, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for merge_info in merge_cells:
            row = merge_info.get('row', 0) + 1
            col = merge_info.get('col', 0) + 1
            rowspan = merge_info.get('rowspan', 1)
            colspan = merge_info.get('colspan', 1)
            
            if rowspan > 1 or colspan > 1:
                start_cell = get_column_letter(col) + str(row)
                end_row = row + rowspan - 1
                end_col = col + colspan - 1
                end_cell = get_column_letter(end_col) + str(end_row)
                
                try:
                    ws.merge_cells(f'{start_cell}:{end_cell}')
                except Exception:
                    pass
        
        for col_idx in range(1, len(data[0]) + 1 if data else 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        wb.save(output_path)
    
    def save_html(self, table_data, output_path):
        data = table_data.get('data', [])
        merge_cells = table_data.get('mergeCells', [])
        
        # html_lines = ['<!DOCTYPE html>', '<html>', '<head>', '<meta charset="UTF-8">', 
        #               '<style>', 
        #               'table { border-collapse: collapse; width: 100%; }',
        #               'td, th { border: 1px solid black; padding: 8px; text-align: center; }',
        #               '</style>', 
        #               '</head>', '<body>', '<table>']
        html_lines=['<table>']
        merge_map = {}
        for merge_info in merge_cells:
            row = merge_info.get('row', 0)
            col = merge_info.get('col', 0)
            rowspan = merge_info.get('rowspan', 1)
            colspan = merge_info.get('colspan', 1)
            
            for r in range(row, row + rowspan):
                for c in range(col, col + colspan):
                    if r == row and c == col:
                        continue
                    merge_map[(r, c)] = 'skip'
            
            merge_map[(row, col)] = {'rowspan': rowspan, 'colspan': colspan}
        
        for row_idx, row in enumerate(data):
            html_lines.append('<tr>')
            for col_idx, value in enumerate(row):
                if (row_idx, col_idx) in merge_map:
                    if merge_map[(row_idx, col_idx)] == 'skip':
                        continue
                    merge_attr = merge_map[(row_idx, col_idx)]
                    rowspan_attr = f' rowspan="{merge_attr["rowspan"]}"' if merge_attr["rowspan"] > 1 else ''
                    colspan_attr = f' colspan="{merge_attr["colspan"]}"' if merge_attr["colspan"] > 1 else ''
                    html_lines.append(f'<td{rowspan_attr}{colspan_attr}>{value}</td>')
                else:
                    html_lines.append(f'<td>{value}</td>')
            html_lines.append('</tr>')
        
        # html_lines.extend(['</table>', '</body>', '</html>'])
        html_lines.append('</table>')
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(html_lines))
